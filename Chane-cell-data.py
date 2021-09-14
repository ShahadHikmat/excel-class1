# give you access to openpyxl and imports workbook
from openpyxl import workbook, load_workbook
# this commmand will the excel sheet from file sources and create a variable to use the file source
workbookVariable = load_workbook('Grades.xlsx')
# create a variable to access the active worksheet
worksheetVarible = workbookVariable.active
# changes the value of the desired cell into another value
worksheetVarible['A2'] = "Hat"
# saves changes to the file in the case value
workbookVariable.save('Grades.xlsx')

# stop here
from openpyxl import workbook, load_workbook
workbookVariable = load_workbook('Grades.xlsx')
worksheetVarible = workbookVariable.active
print(workbookVariable.sheetnames)

# stop

from openpyxl import workbook, load_workbook
workbookVariable = load_workbook('Grades.xlsx')
workbookVariable.create_chartsheet('new-sheet')
print(workbookVariable.sheetnames)






